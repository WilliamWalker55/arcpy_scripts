#!/usr/bin/python
# -*- coding: utf-8 -*-
import arcpy
import arcpy.mapping as mapping
import glob
import os
from openpyxl import Workbook

#Name of folder containing mxd files
folder = "TEST"
path = "Z:\\mxd\\" + folder


#Excel doc path 
out = "Z:/mxd/Dokumentation/" + folder + "_data_source.xlsx"

wb= Workbook()
sheetCount = 1 #count for each mxd
allCount = 3 #count for first sheet including all sources

#Sheet and column names for all data sources (first sheet)
ws = wb.create_sheet("Alla",0)
ws['A1'] = "Source"
ws['B1'] = "mxd"


for file in os.listdir(path):
    if file.endswith(".mxd"):
        print file   
        rowCount = 2
            
        mxd_path = os.path.join(path,file)
        mxd = mapping.MapDocument(mxd_path)
            
        name = file.split('.')[0]
    
        #Sheet names cannot be more than 31 characters
        name = name[:29]
        
        print name

        #Column names for each mxd   
        ws1 = wb.create_sheet(name,sheetCount)
        ws1['A1'] = "Layer name"
        ws1['B1'] = "Layer Type"
        ws1['C1'] = "Source"
           
           
        for lyr in arcpy.mapping.ListLayers(mxd):
            # List feature classes
            if lyr.supports("DATASOURCE")and lyr.supports("datasetName"):
                layer = lyr.dataSource.encode('utf-8')
                print layer
                layer_name = lyr.name.encode('utf-8')
                print layer_name
                ws['A' + str(allCount)] = layer
                ws['B' + str(allCount)] = name
                ws1['A' + str(rowCount)] = layer_name
                ws1['C' + str(rowCount)] = layer
                ws1['B' + str(rowCount)] = "Feature class"
                rowCount = rowCount + 1
                allCount = allCount + 1
                    
            # List WMS services   
            elif lyr.supports("SERVICEPROPERTIES"):
                servProp = lyr.serviceProperties
                layer_name = lyr.name.encode('utf-8')
                print layer_name
                ws1['A' + str(rowCount)] = layer_name
                ws1['B' + str(rowCount)] = servProp.get('ServiceType', 'N/A')
                ws1['C' + str(rowCount)] = servProp.get('URL', 'N/A')
                rowCount = rowCount + 1

        sheetCount = sheetCount + 1
        del mxd
          
           
wb.save(out)
print "Script complete"
        

